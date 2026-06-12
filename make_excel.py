# -*- coding: utf-8 -*-
"""
松江市エリアブランディング戦略策定 タスク管理用Excelを生成する。
ガントチャートWebアプリ（task_gantt.html）にインポートできる列構成で、
見やすい書式（色分け・罫線・データ入力規則など）を付与する。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\takeshita\Documents\松江観光\タスク管理\松江タスク管理.xlsx"

# 列定義（ヘッダー名, 幅）
COLUMNS = [
    ("大項目", 16),
    ("中項目", 16),
    ("タスク名", 28),
    ("担当組織", 14),
    ("担当者名", 12),
    ("ステータス", 12),
    ("開始日", 13),
    ("終了日", 13),
    ("打合せ", 13),
    ("提出", 13),
    ("承認", 13),
    ("備考", 26),
]

# タスクデータ（大項目, 中項目, タスク名, 担当組織）
# 担当組織: 空=エブリプラン / G=グレイスデザイン / 2=両者 / 承認=観光協会
TASKS = [
    ("1-1 チラシ", "文面", "エブリプラン"),
    ("1-1 チラシ", "デザイン", "グレイスデザイン"),
    ("1-1 チラシ", "受付メアド", "エブリプラン"),
    ("1-1 チラシ", "障がい者・子連れ対応", "エブリプラン"),
    ("1-1 チラシ", "駐車場確認", "エブリプラン"),
    ("1-1 チラシ", "承認", "観光協会"),
    ("1-2 企画", "プログラム", "エブリプラン"),
    ("1-2 企画", "会場デザイン", "グレイスデザイン"),
    ("1-2 企画", "喫茶コーナー", "エブリプラン"),
    ("1-2 企画", "承認", "観光協会"),
    ("募集リスト", "記入", "両者"),
    ("募集リスト", "提出", "エブリプラン"),
    ("募集リスト", "順位付け", "エブリプラン"),
    ("募集リスト", "再確認", "エブリプラン"),
    ("募集リスト", "承認", "観光協会"),
    ("会場", "正式申込", "エブリプラン"),
    ("会場", "下見", "エブリプラン"),
    ("会場", "設備手配（会場に無い物）", "エブリプラン"),
    ("会場", "備品購入", "エブリプラン"),
    ("会場", "承認", "観光協会"),
    ("講演", "講師打診", "エブリプラン"),
    ("講演", "宿泊予約", "エブリプラン"),
    ("講演", "ロジ検討", "エブリプラン"),
    ("講演", "承認", "観光協会"),
    ("講演", "プログラム連絡", "エブリプラン"),
    ("講演", "MTG", "エブリプラン"),
    ("募集", "リスト順に訪問", "両者"),
    ("募集", "中間集計・報告", "エブリプラン"),
    ("募集", "リストに追加", "両者"),
    ("募集", "締め切り前集計・報告", "エブリプラン"),
    ("募集", "返信", "エブリプラン"),
    ("メディア", "プレスリリース文案作成", "エブリプラン"),
    ("メディア", "承認・リリース", "観光協会"),
    ("メディア", "メディア事前打合せ", "エブリプラン"),
]
BIG_CAT = "1. 現状調査WS"

# ── 配色 ──
HEADER_FILL = PatternFill("solid", fgColor="00695C")
HEADER_FONT = Font(name="Meiryo", bold=True, color="FFFFFF", size=11)

ORG_FILL = {
    "エブリプラン": PatternFill("solid", fgColor="E8EAF6"),
    "グレイスデザイン": PatternFill("solid", fgColor="FCE4EC"),
    "両者": PatternFill("solid", fgColor="F3E5F5"),
    "観光協会": PatternFill("solid", fgColor="E0F2F1"),
}
# 中項目グループの交互背景
SUBCAT_FILL_A = PatternFill("solid", fgColor="FFFFFF")
SUBCAT_FILL_B = PatternFill("solid", fgColor="F7F9FB")

THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BASE_FONT = Font(name="Meiryo", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "タスク一覧"

# タイトル行
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
tcell = ws.cell(row=1, column=1, value="松江市エリアブランディング戦略策定　タスク管理")
tcell.font = Font(name="Meiryo", bold=True, size=14, color="00695C")
tcell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 26

# ヘッダー行（2行目）
HEADER_ROW = 2
for ci, (name, width) in enumerate(COLUMNS, start=1):
    c = ws.cell(row=HEADER_ROW, column=ci, value=name)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER
    ws.column_dimensions[get_column_letter(ci)].width = width
ws.row_dimensions[HEADER_ROW].height = 22

# データ行
col_idx = {name: i + 1 for i, (name, _) in enumerate(COLUMNS)}
start = HEADER_ROW + 1
subcat_order = []
for sub, _, _ in TASKS:
    if sub not in subcat_order:
        subcat_order.append(sub)

for ri, (sub, task, org) in enumerate(TASKS):
    row = start + ri
    values = {
        "大項目": BIG_CAT,
        "中項目": sub,
        "タスク名": task,
        "担当組織": org,
        "担当者名": "",
        "ステータス": "未着手",
        "開始日": "",
        "終了日": "",
        "打合せ": "",
        "提出": "",
        "承認": "",
        "備考": "",
    }
    # 中項目グループで交互背景
    group_idx = subcat_order.index(sub)
    band = SUBCAT_FILL_A if group_idx % 2 == 0 else SUBCAT_FILL_B

    for name, ci in col_idx.items():
        c = ws.cell(row=row, column=ci, value=values[name])
        c.font = BASE_FONT
        c.border = BORDER
        if name in ("タスク名", "中項目", "大項目", "備考"):
            c.alignment = LEFT
        else:
            c.alignment = CENTER
        c.fill = band
    # 担当組織セルだけ組織色で上書き
    org_cell = ws.cell(row=row, column=col_idx["担当組織"])
    if org in ORG_FILL:
        org_cell.fill = ORG_FILL[org]
    # 日付列の表示形式
    for dname in ("開始日", "終了日", "打合せ", "提出", "承認"):
        ws.cell(row=row, column=col_idx[dname]).number_format = "yyyy/m/d"
    ws.row_dimensions[row].height = 20

last_row = start + len(TASKS) - 1

# ── 入力規則（ドロップダウン） ──
# ステータス
dv_status = DataValidation(type="list", formula1='"未着手,進行中,完了,保留"', allow_blank=True)
ws.add_data_validation(dv_status)
dv_status.add(f"{get_column_letter(col_idx['ステータス'])}{start}:{get_column_letter(col_idx['ステータス'])}{last_row}")
# 担当組織
dv_org = DataValidation(type="list", formula1='"エブリプラン,グレイスデザイン,両者,観光協会"', allow_blank=True)
ws.add_data_validation(dv_org)
dv_org.add(f"{get_column_letter(col_idx['担当組織'])}{start}:{get_column_letter(col_idx['担当組織'])}{last_row}")

# ウィンドウ枠の固定（ヘッダーまで＝3行目以降スクロール／タスク名まで固定）
ws.freeze_panes = ws.cell(row=start, column=col_idx["担当組織"])

# オートフィルター
ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(COLUMNS))}{last_row}"

# ── 凡例シート ──
ws2 = wb.create_sheet("凡例・入力ガイド")
guide = [
    ("項目", "説明・入力例"),
    ("大項目", "フェーズ。ガントチャートのグループ見出しになる（例：1. 現状調査WS）"),
    ("中項目", "詳細グループ（例：1-1 チラシ、会場、講演 など）"),
    ("タスク名", "個別の作業名"),
    ("担当組織", "エブリプラン / グレイスデザイン / 両者 / 観光協会 から選択"),
    ("担当者名", "実際の作業者名（例：竹下）"),
    ("ステータス", "未着手 / 進行中 / 完了 / 保留 から選択"),
    ("開始日", "タスク開始予定日（例：2026/6/15）→ ガントのバー始点"),
    ("終了日", "タスク終了予定日（例：2026/6/25）→ ガントのバー終点"),
    ("打合せ", "打合せ・会議の日付（●青で表示）"),
    ("提出", "資料・成果物の提出日（◆赤で表示）"),
    ("承認", "発注者（観光協会）の承認日（◎緑で表示）"),
    ("備考", "メモ。ガントチャートのツールチップに表示"),
]
for ri, (a, b) in enumerate(guide, start=1):
    ca = ws2.cell(row=ri, column=1, value=a)
    cb = ws2.cell(row=ri, column=2, value=b)
    ca.border = BORDER
    cb.border = BORDER
    ca.alignment = Alignment(vertical="center")
    cb.alignment = Alignment(vertical="center", wrap_text=True)
    if ri == 1:
        ca.fill = HEADER_FILL; ca.font = HEADER_FONT; ca.alignment = CENTER
        cb.fill = HEADER_FILL; cb.font = HEADER_FONT; cb.alignment = CENTER
    else:
        ca.font = Font(name="Meiryo", size=10, bold=True)
        cb.font = BASE_FONT
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 64

wb.save(OUT)
print("saved:", OUT)
print("rows:", len(TASKS))
